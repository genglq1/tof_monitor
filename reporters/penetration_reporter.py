#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
穿透报告报表（复刻旧版 run_penetration.py 的写出逻辑）。

- 工作表名：穿透固收&权益
- 文件名：{项目代码}_{信托名称}_{日期}_穿透报告.xlsx
- 颜色预警：穿透固收占比 / 穿透权益占比 超过 80 红色、超过 70 黄色
- 列宽：按内容自适应（上限 50）
"""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class PenetrationReporter:
    SHEET_NAME = "穿透固收&权益"

    def write_report(self, result, output_dir: Path) -> Path:
        """将 PenetrationResult 写出为穿透报告 Excel，返回文件路径。"""
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        output_file = output_dir / f"{result.code}_{result.trust_name}_{result.query_date}_穿透报告.xlsx"
        result.df.to_excel(output_file, sheet_name=self.SHEET_NAME, index=False)

        # 颜色预警（复刻旧 run_penetration.generate_report 末尾样式处理）
        try:
            wb = load_workbook(output_file)
            ws = wb[self.SHEET_NAME]
            for col_idx, val in [(7, result.total_fixed), (8, result.total_equity)]:
                cell = ws.cell(row=2, column=col_idx)
                if cell.value:
                    try:
                        v = float(cell.value)
                        if v > 80:
                            cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                        elif v > 70:
                            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    except Exception:
                        pass
            for col in ws.columns:
                max_len = max(len(str(cell.value)) for cell in col if cell.value) + 2
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 50)
            wb.save(output_file)
        except Exception:
            pass

        return output_file
