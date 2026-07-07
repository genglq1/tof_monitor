#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
估值统计报表生成器

复刻旧版 run_valuation.run 的 Excel 写出逻辑：统计结果工作表 + 空表格记录工作表，
列宽、表头填充/字体/居中样式与原脚本一致。
"""

from pathlib import Path
import pandas as pd

from parsers.valuation_stats import COLUMNS_ORDER, ValuationStat


class ValuationReporter:
    def write_stats(self, stats, output_file, empty_files=None):
        """写出 信托产品估值统计结果.xlsx。

        stats: list[ValuationStat]
        empty_files: list[dict]，可选，非空时追加「空表格记录」工作表
        """
        output_file = Path(output_file)
        if not stats:
            print("没有数据可导出")
            return

        rows = [s.to_row() for s in stats]
        df = pd.DataFrame(rows)
        df = df[[c for c in COLUMNS_ORDER if c in df.columns]]

        output_file.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='统计结果', index=False)
            ws = writer.sheets['统计结果']
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 12

            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_alignment = Alignment(horizontal='center', vertical='center')
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            if empty_files:
                df_empty = pd.DataFrame(empty_files)
                df_empty.to_excel(writer, sheet_name='空表格记录', index=False)

        print(f"结果已导出到: {output_file}")
