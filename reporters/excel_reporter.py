from pathlib import Path
from typing import List
from decimal import Decimal
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from core.models import AssetOverview, HoldingDetail


class ExcelReporter:
    HEADER_COLOR = "4472C4"   # 蓝色表头
    WARNING_COLOR = "FF6B6B"  # 红色警告
    HIGHLIGHT_COLOR = "FFFF00"  # 黄色高亮

    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _format_worksheet(self, ws, col_widths: dict = None, freeze: str = 'A2'):
        """通用格式化：列宽、自动换行、冻结"""
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width
        if freeze:
            ws.freeze_panes = freeze

    def _style_header(self, ws, color: str = None):
        """表头样式：蓝色背景、白色粗体、居中"""
        color = color or self.HEADER_COLOR
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        font = Font(bold=True, color="FFFFFF", size=11)
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment

    def _center_data_rows(self, ws, start_row: int = 2):
        """数据行居中"""
        alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows(min_row=start_row):
            for cell in row:
                cell.alignment = alignment

    def report_overviews(self, data: List[AssetOverview], date_str: str) -> Path:
        rows = []
        for d in data:
            rows.append({
                '项目代码': d.project_code,
                '项目名称': d.project_name,
                '估值日期': str(d.valuation_date),
                '银行存款_市值': self._fmt_number(d.bank_deposit.market_value),
                '银行存款_占净值%': self._fmt_pct(d.bank_deposit.nav_pct),
                '信托保障基金_市值': self._fmt_number(d.guarantee_fund.market_value),
                '信托保障基金_占净值%': self._fmt_pct(d.guarantee_fund.nav_pct),
                '证券投资合计_市值': self._fmt_number(d.securities_total.market_value),
                '证券投资合计_占净值%': self._fmt_pct(d.securities_total.nav_pct),
                '实收信托_市值': self._fmt_number(d.paid_in_trust.market_value),
                '实收信托_占净值%': self._fmt_pct(d.paid_in_trust.nav_pct),
                '负债类合计_市值': self._fmt_number(d.liabilities_total.market_value),
                '负债类合计_占净值%': self._fmt_pct(d.liabilities_total.nav_pct),
                '资产类合计_市值': self._fmt_number(d.assets_total.market_value),
                '资产类合计_占净值%': self._fmt_pct(d.assets_total.nav_pct),
            })

        df = pd.DataFrame(rows)
        path = self.output_dir / f"资产总览_{date_str}.xlsx"

        col_widths = {
            'A': 12, 'B': 35, 'C': 12,
            'D': 15, 'E': 15,
            'F': 18, 'G': 18,
            'H': 18, 'I': 18,
            'J': 15, 'K': 15,
            'L': 15, 'M': 15,
            'N': 15, 'O': 15,
        }

        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='统计结果', index=False)
            ws = writer.sheets['统计结果']
            self._style_header(ws)
            self._center_data_rows(ws)
            self._format_worksheet(ws, col_widths, freeze='A2')

        return path

    def report_holdings(self, data: List[HoldingDetail], date_str: str) -> Path:
        rows = [{
            '项目代码': h.project_code,
            '项目名称': h.project_name,
            '投资标的': h.target_name,
            '科目代码': h.account_code,
            '类型': h.holding_type.value,
            '市值': self._fmt_number(h.market_value),
            '占比%': self._fmt_pct(h.nav_pct),
        } for h in data]

        df = pd.DataFrame(rows)
        path = self.output_dir / f"持仓明细_{date_str}.xlsx"

        col_widths = {
            'A': 12, 'B': 35, 'C': 40,
            'D': 12, 'E': 10,
            'F': 18, 'G': 10,
        }

        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='持仓明细', index=False)
            ws = writer.sheets['持仓明细']
            self._style_header(ws)
            self._center_data_rows(ws)
            self._format_worksheet(ws, col_widths, freeze='A2')

        return path

    def _fmt_number(self, val) -> float:
        try:
            if isinstance(val, Decimal):
                return float(val)
            return float(val) if val else 0.0
        except:
            return 0.0

    def _fmt_pct(self, val) -> float:
        try:
            if isinstance(val, Decimal):
                return float(val)
            return float(val) if val else 0.0
        except:
            return 0.0
